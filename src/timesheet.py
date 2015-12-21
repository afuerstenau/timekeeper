from openpyxl import load_workbook
import datetime


class Timesheet:
    def __init__(self, filename):
        self.filename = filename
        self.book = load_workbook(self.filename)
        self.sheet = self.book.get_sheet_by_name("Zeiten")
        self.sum_sheet = self.book.get_sheet_by_name("Summen")
        self.normative_working_time_in_hours = self.sum_sheet.cell("A2").value

    def append_workentry(self, workentry):
        current_date = datetime.date.today()-datetime.timedelta(1)
        self.sheet.append([current_date.strftime("%d.%m.%Y"), workentry.starttime(), workentry.endtime(), workentry.description(), workentry.lunchbreak(), workentry.working_time()])
        sum_working_time = self.calculate_sum_working_time(workentry)
        self.set_sum_working_time(sum_working_time)
        self.book.save(self.filename)
        return sum_working_time

    def calculate_sum_working_time(self, workentry = None):
        sum = 0
        if workentry != None:
            sum = workentry.working_time()
        i=2
        while i < self.sheet.max_row:
            value = self.sheet.cell(row=i, column=6).value
            try:
                float_value = float(value)
                sum = sum +float_value
            except Exception as e:
                pass
            i+=1
        return sum

    def set_sum_working_time(self, sum):
        self.sheet = self.book.get_sheet_by_name("Summen")
        self.sheet.cell("B2").value = sum
        self.book.save(self.filename)

    def get_sum_working_time(self):
        return self.calculate_sum_working_time()

    def get_diff_normative_actual_working_time_in_hours(self):
        return self.normative_working_time_in_hours-self.calculate_sum_working_time()

    def get_remaining_days(self):
        return self.get_diff_normative_actual_working_time_in_hours()/8
